"""Report class
"""
from typing import List
import json
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook


class Report():
    """Report class to generate XLSX reports
    """
    def __init__(self):
        """Init Report class
        """
        self.workbook = None
        self.headers = []
        self.data = []
        self.alphabet = [chr(a) for a in range(ord("A"), ord("Z")+1)]
        self.config = {
            "bold": Font(bold=True)
        }
        with open("Report/settings.json", "r", encoding="utf-8") as f:
            self.settings = json.load(f)

    def __parse_number(self, number: str) -> int | float | str:
        """internal: Set correct type for int/float after reading from file.

        Args:
            number (str): Number as string

        Returns:
            int|float|str: Number with correct type; str if no conversion possible
        """
        try:
            return int(number)
        except ValueError:
            pass
        try:
            return float(number)
        except Exception:
            return number

    def set_headers(self, headers: dict):
        """Set headers for each column

        Args:
            headers (dict): Column headers
        """
        self.headers = headers

    def set_data(self, data: List[list]):
        """Data for each row

        Args:
            data (List[list]): Nested list with data for each row
        """
        self.data = data

    def save(self, fname: str):
        """Save spreadsheet

        Args:
            fname (str): Filename of spreadsheet

        Raises:
            Exception: Too many header values; may be fixed in future
        """
        self.workbook = Workbook()
        sheet = self.workbook.active
        pos = 1
        try:
            head = [f"{c}{pos}" for c in self.alphabet[:len(self.headers)]]
        except Exception as exception:
            # TODO: Fix issue with spreadsheet coordinates
            raise exception from Exception("Too many header values")
        for cell, val in zip(head, self.headers):
            sheet[cell] = val
            sheet[cell].font = self.config["bold"]
            # TODO: Find out how to freeze y-axis, not x-axis
            # sheet.freeze_panes = cell

        # Fix for above TODO
        cell = sheet['B2']
        sheet.freeze_panes = cell

        for row in self.data:
            print(row)
            pos += 1
            first = True
            head = [f"{c}{pos}" for c in self.alphabet[:len(self.headers)]]
            for i, cell, val in zip(range(len(row)), head, row):
                sheet[cell] = val
                if first:
                    sheet[cell].font = self.config["bold"]
                    first = False
                else:
                    if val < self.settings[self.headers[i]]["threshold"]:
                        color = self.settings[self.headers[i]]["lower"]
                        sheet[cell].fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid")
                    else:
                        color = self.settings[self.headers[i]]["higher"]
                        sheet[cell].fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid")

        # Add filter to sort results
        sheet.auto_filter.ref = sheet.dimensions

        self.workbook.save(fname)

    def as_csv(self, fname: str, sep: str = ","):
        """Save data as CSV

        Args:
            fname (str): Filename of CSV file
            sep (str, optional): Value separator. Defaults to ",".
        """
        with open(fname, 'w', encoding="utf-8") as f:
            f.write(sep.join(self.headers) + "\n")
            for row in self.data:
                f.write(','.join([str(val) for val in row]) + "\n")

    def from_csv(self, fname: str, xlsx_file: str, sep: str = ","):
        """Create report from CSV

        Args:
            fname (str): Path to CSV file
            sep (str, optional): Value separator. Defaults to ",".
        """
        try:
            with open(fname, 'r', encoding="utf-8") as f:
                data = f.read().splitlines()
        except Exception as exception:
            raise exception from FileNotFoundError(f"Could not open file '{fname}'")
        headers = data[0].split(",")
        values = [[self.__parse_number(n) for n in row.split(sep)]
                  for row in data[1:]]
        self.set_headers(headers)
        self.set_data(values)
        self.save(xlsx_file)
